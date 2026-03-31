import os
import re
import tkinter as tk
from copy import copy
from tkinter import filedialog, messagebox, simpledialog

import pdfplumber
from openpyxl import load_workbook
from tkinterdnd2 import DND_FILES, TkinterDnD


class QCFApp:
    VENDOR_KEYWORDS = [
        ("UMEDIC", "UMEDIC"),
        ("AZONIC", "AZONIC"),
        ("CERT ACADEMY", "CERT ACADEMY"),
    ]

    COMPANY_SUFFIXES = [
        r"\bSDN\.?\s*BHD\.?\b",
        r"\bBERHAD\b",
        r"\bBHD\.?\b",
        r"\bGROUP\b",
        r"\bHOLDINGS?\b",
        r"\bRESOURCES?\b",
        r"\bENGINEERING\b",
        r"\bTECHNOLOG(?:Y|IES)\b",
        r"\bSOLUTIONS?\b",
        r"\bCONSULT(?:ING|ANCY)\b",
        r"\bSERVICES?\b",
        r"\bTRADING\b",
        r"\bENTERPRISE\b",
        r"\bACADEMY\b",
        r"\bINSTITUTE\b",
    ]

    SKIP_MARKERS = [
        "LEADER", "ATTN", "ATTENTION", "TO:", "TO ", "DATE", "QUOTATION",
        "QUOTE", "REF", "OUR REF", "PURCHASE", "TEL", "FAX", "EMAIL",
        "WWW.", "HTTP", "ADDRESS", "ADDR", "COMPANY :", "NAME OF COMPANY",
    ]

    ADDRESS_MARKERS = [
        "JALAN", "LOT", "NO.", "NO ", "ROAD", "KAMPUNG", "TAMAN",
        "PERSIARAN", "BANDAR", "PENANG", "KEDAH",
    ]

    FOOTER_LABELS = {
        "payment term",
        "delivery period",
        "quotation reference no.",
        "total value excluded tax",
        "total value included tax",
        "sales & service tax",
        "price comparison between lowest value",
        "price comparison between lowest %",
    }

    def __init__(self, root):
        self.root = root
        self.root.title("Quotation to QCF Converter - Production Build")
        self.root.geometry("550x450")

        self.template_path = ""
        self.pdf_list = []

        # --- UI Elements ---
        tk.Label(root, text="Step 1: Select QCF Template", font=("Arial", 10, "bold")).pack(pady=5)
        self.btn_template = tk.Button(root, text="Browse Excel Template", command=self.load_template)
        self.btn_template.pack()

        tk.Label(root, text="Step 2: Drag & Drop Quotations Below", font=("Arial", 10, "bold")).pack(pady=15)
        self.drop_target = tk.Listbox(root, width=70, height=10)
        self.drop_target.pack(padx=20, pady=5)

        self.drop_target.drop_target_register(DND_FILES)
        self.drop_target.dnd_bind("<<Drop>>", self.handle_drop)

        self.btn_add = tk.Button(root, text="Add PDFs", command=self.add_files)
        self.btn_add.pack(pady=5)

        self.btn_run = tk.Button(
            root,
            text="Process & Generate QCF",
            command=self.start_processing,
            bg="green",
            fg="white",
            font=("Arial", 12, "bold"),
        )
        self.btn_run.pack(pady=20)

    # ---------------- UI helpers ----------------
    def load_template(self):
        self.template_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.template_path:
            self.btn_template.config(text=os.path.basename(self.template_path), fg="blue")

    def handle_drop(self, event):
        files = self.root.tk.splitlist(event.data)
        self._add_pdf_files(files)

    def add_files(self):
        files = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        if files:
            self._add_pdf_files(files)

    def _add_pdf_files(self, files):
        for f in files:
            if f.lower().endswith(".pdf") and f not in self.pdf_list:
                self.pdf_list.append(f)
                self.drop_target.insert(tk.END, os.path.basename(f))

    # ---------------- Processing ----------------
    def start_processing(self):
        if not self.template_path or not self.pdf_list:
            messagebox.showwarning("Missing Files", "Please provide the template and at least one PDF.")
            return

        item_capacity = simpledialog.askinteger(
            "Item Rows",
            "How many item rows are available from Row 8 onwards?",
            minvalue=1,
            parent=self.root,
        )
        if not item_capacity:
            messagebox.showwarning("Missing Input", "Please enter the number of item rows in the template.")
            return

        output_path = filedialog.asksaveasfilename(
            title="Save Output As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Consolidated_QCF_Result.xlsx",
        )
        if not output_path:
            return

        success_count = 0
        wb = load_workbook(self.template_path)
        ws = wb.active

        for pdf_path in self.pdf_list:
            try:
                self.process_single_pdf(pdf_path, ws, item_capacity)
                success_count += 1
            except Exception as e:
                print(f"Error processing {os.path.basename(pdf_path)}: {e}")

        if success_count > 0:
            wb.save(output_path)
            messagebox.showinfo("Success", f"Data inserted!\nSaved as: {os.path.basename(output_path)}")
            try:
                os.startfile(output_path)
            except Exception as exc:
                messagebox.showwarning("Open Failed", f"Saved, but could not open file:\n{exc}")
        else:
            messagebox.showerror("Error", "Failed to process the files.")

    def process_single_pdf(self, pdf_path, ws, item_capacity):
        extracted = self.extract_pdf_data(pdf_path)
        print(f"Extracted from {os.path.basename(pdf_path)}:", extracted)

        rate_col, amt_col = self.get_vendor_columns(ws, extracted["vendor"])

        start_row = 8
        footer_start = self.find_footer_start(ws)
        items_count = len(extracted["items"])

        # Find first empty row in Description column between start_row and footer
        empty_row = None
        for r in range(start_row, footer_start):
            if not ws.cell(row=r, column=3).value:
                empty_row = r
                break
        start_row = empty_row if empty_row is not None else footer_start

        # Ensure there are enough rows before footer for all items
        if start_row == 8 and footer_start is not None:
            rows_to_insert = max(0, items_count - item_capacity)
            if rows_to_insert > 0:
                ws.insert_rows(footer_start, amount=rows_to_insert)
                footer_start += rows_to_insert
                self.copy_item_row_styles(ws, footer_start - rows_to_insert, rows_to_insert)
        else:
            available_rows = footer_start - start_row
            rows_to_insert = max(0, items_count - available_rows)
            if rows_to_insert > 0:
                ws.insert_rows(footer_start, amount=rows_to_insert)
                footer_start += rows_to_insert
                self.copy_item_row_styles(ws, footer_start - rows_to_insert, rows_to_insert)
            elif start_row >= footer_start:
                ws.insert_rows(footer_start, amount=items_count)
                footer_start += items_count
                self.copy_item_row_styles(ws, footer_start - items_count, items_count)

        # Write items
        for i, item in enumerate(extracted["items"]):
            curr_row = start_row + i
            desc, qty, uom, rate = item
            ws.cell(row=curr_row, column=3).value = desc
            ws.cell(row=curr_row, column=6).value = qty
            ws.cell(row=curr_row, column=7).value = uom
            ws.cell(row=curr_row, column=rate_col).value = rate
            ws.cell(row=curr_row, column=amt_col).value = qty * rate

        # Write footer terms based on labels
        pay_row = self.find_label_row(ws, "payment term")
        del_row = self.find_label_row(ws, "delivery period")
        ref_row = self.find_label_row(ws, "quotation reference no.")

        if pay_row:
            ws.cell(row=pay_row, column=rate_col).value = extracted["payment"]
        if del_row:
            ws.cell(row=del_row, column=rate_col).value = extracted["delivery"]
        if ref_row:
            ws.cell(row=ref_row, column=rate_col).value = extracted["ref_no"]

    # ---------------- Extraction ----------------
    def extract_pdf_data(self, pdf_path):
        data = {
            "vendor": "Unknown Company",
            "ref_no": "N/A",
            "payment": "N/A",
            "delivery": "N/A",
            "items": [],
        }

        with pdfplumber.open(pdf_path) as pdf:
            text, lines = self.get_text_and_lines(pdf)

            data["vendor"] = self.detect_vendor(text, lines)
            data["ref_no"] = self.extract_ref(lines, text)

            payment, delivery = self.extract_terms(text)
            data["payment"], data["delivery"] = self.normalize_terms(payment, delivery)

            items = self.extract_items_from_lines(lines)
            if not items:
                items = self.extract_items_from_tables(pdf)
            if not items:
                items = [self.fallback_total_item(text)]
            data["items"] = items

        return data

    def get_text_and_lines(self, pdf):
        pages_text = []
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                pages_text.append(page_text)
        text = "\n".join(pages_text)
        lines = [line.strip() for line in text.split("\n") if line.strip()]
        return text, lines

    def detect_vendor(self, text, lines):
        text_upper = text.upper()
        for key, name in self.VENDOR_KEYWORDS:
            if key in text_upper:
                return name

        best_line = None
        best_score = -1
        for line in lines[:20]:
            upper = line.upper()
            if any(m in upper for m in self.SKIP_MARKERS):
                continue

            score = 0
            if any(m in upper for m in self.ADDRESS_MARKERS):
                score -= 2
            if re.search(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", line):
                score -= 3
            if re.search(r"\b\d{1,2}:\d{2}\s*(AM|PM)?\b", upper):
                score -= 2
            for suf in self.COMPANY_SUFFIXES:
                if re.search(suf, upper):
                    score += 5
            if upper.isupper():
                score += 1

            if score > best_score:
                best_score = score
                best_line = line

        if best_line and best_score > 0:
            return best_line[:60]

        for line in lines[:20]:
            upper = line.upper()
            if "LEADER" in upper:
                continue
            if "HTTP" in upper or "WWW." in upper:
                continue
            if "EMAIL" in upper or "TEL" in upper or "FAX" in upper:
                continue
            return line[:60]

        return "Unknown Company"

    def extract_ref(self, lines, text):
        def clean_ref(raw):
            val = raw.strip()
            val = re.split(
                r"\b(Date|Quotation Date|Valid Through|Payment Term|Attention|Attn)\b",
                val,
                flags=re.IGNORECASE,
            )[0].strip()
            return val.rstrip(" .,:;")

        for line in lines:
            for pat in [
                r"\bOur\s*Ref\s*[:\-]\s*(.+)$",
                r"\bOurRef\s*[:\-]\s*(.+)$",
                r"\bReference\s*Number\s*[:\-]\s*(.+)$",
                r"\bRef(?:erence)?\s*(?:No\.?|#|Number)?\s*[:\-]\s*(.+)$",
                r"\b(?:Quotation|Quote)\s*(?:No\.?|#|Ref\.?|Reference)?\s*[:\-]\s*(.+)$",
            ]:
                m = re.search(pat, line, re.IGNORECASE)
                if m:
                    return clean_ref(m.group(1))

        ref_patterns = [
            r"\bOur\s*Ref\s*[:\-]\s*([A-Z0-9][A-Z0-9\-\/&_.]+)",
            r"\bOurRef\s*[:\-]\s*([A-Z0-9][A-Z0-9\-\/&_.]+)",
            r"\bQuotation\s*No\.?\s*[:\-]\s*([A-Z0-9][A-Z0-9\-\/&_.]+)",
            r"\bQuote\s*No\.?\s*[:\-]\s*([A-Z0-9][A-Z0-9\-\/&_.]+)",
            r"\bRef\.?\s*(?:No\.?|#|Number)?\s*[:\-]\s*([A-Z0-9][A-Z0-9\-\/&_.]+)",
            r"\bReference\s*Number\s*[:\-]\s*([A-Z0-9][A-Z0-9\-\/&_.]+)",
        ]
        ref_candidates = []
        for pat in ref_patterns:
            ref_candidates.extend(re.findall(pat, text, re.IGNORECASE))
        if ref_candidates:
            return max(ref_candidates, key=len)
        return "N/A"

    def extract_terms(self, text):
        payment = "N/A"
        delivery = "N/A"
        pay_match = re.search(
            r"(?:Payment|Terms?)[:\s]*(.*?(?:Days|COD|Cash|Advance|Month).*?)(?=\n|$)",
            text,
            re.IGNORECASE,
        )
        if pay_match:
            payment = pay_match.group(1).strip()

        del_match = re.search(
            r"(?:Delivery|Lead\s*Time)[:\s]*(.*?(?:Weeks|Days|Months|ARO).*?)(?=\n|$)",
            text,
            re.IGNORECASE,
        )
        if del_match:
            delivery = del_match.group(1).strip()

        return payment, delivery

    def normalize_terms(self, payment, delivery):
        if payment != "N/A":
            payment = re.sub(
                r"(?i)^(?:term\s*of\s*payment|payment\s*term|payment)\s*[:\-]*\s*",
                "",
                payment,
            ).strip()
        if delivery != "N/A":
            delivery = re.sub(
                r"(?i)^(?:delivery\s*period|delivery|lead\s*time)\s*[:\-]*\s*",
                "",
                delivery,
            ).strip()

        pay_match = re.search(r"(?i)\b(cash|cod|advance|\d+\s*days?)\b", payment)
        if pay_match:
            payment = pay_match.group(0).strip()

        del_match = re.search(r"(?i)\b\d+\s*(?:-\s*\d+\s*)?(weeks?|days?|months?|aro)\b", delivery)
        if del_match:
            delivery = del_match.group(0).strip()

        return payment or "N/A", delivery or "N/A"

    def extract_items_from_lines(self, lines):
        items = []

        def parse_item_line(line):
            if not re.match(r"^\s*\d+(?:\.\d+)?\.?\s+", line):
                return None

            m = re.search(
                r"^(?:\d+(?:\.\d+)?\s+)?(.+?)\s+(\d+(?:\.\d+)?)([A-Za-z]+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$",
                line,
            )
            if m:
                return m.group(1).strip().rstrip("."), float(m.group(2)), m.group(3), float(m.group(4).replace(",", ""))

            m = re.search(
                r"^(?:\d+(?:\.\d+)?\s+)?(.+?)\s+L\/S\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$",
                line,
                re.IGNORECASE,
            )
            if m:
                return m.group(1).strip().rstrip("."), 1.0, "L/S", float(m.group(2).replace(",", ""))

            m = re.search(
                r"^(?:\d+(?:\.\d+)?\s+)?(.+?)\s+(\d+(?:\.\d+)?)([A-Za-z]+)\s+([\d,]+\.\d{2})\s*$",
                line,
            )
            if m:
                return m.group(1).strip().rstrip("."), float(m.group(2)), m.group(3), float(m.group(4).replace(",", ""))

            return None

        in_items = False
        last_idx = None
        skip_next = False
        pending_desc = None

        for i, line in enumerate(lines):
            if skip_next:
                skip_next = False
                continue

            upper = line.upper()
            if "NO" in upper and "DESCRIPTION" in upper and "QTY" in upper:
                in_items = True
                continue
            if "TERMS" in upper or "CONDITION" in upper or "TOTAL" in upper or "SST" in upper:
                in_items = False
                last_idx = None
                continue
            if not in_items:
                continue

            if pending_desc:
                m_qty_only = re.search(r"^\s*(\d+(?:\.\d+)?)([A-Za-z]+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$", line)
                if m_qty_only:
                    qty = float(m_qty_only.group(1))
                    unit = m_qty_only.group(2)
                    rate = float(m_qty_only.group(3).replace(",", ""))
                    items.append((pending_desc, int(qty) if qty.is_integer() else qty, unit, rate))
                    last_idx = len(items) - 1
                    pending_desc = None
                    continue

                m_tail = re.search(
                    r"^(.*)\s+(\d+(?:\.\d+)?)([A-Za-z]+)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$",
                    line,
                )
                if m_tail:
                    extra_desc = m_tail.group(1).strip()
                    qty = float(m_tail.group(2))
                    unit = m_tail.group(3)
                    rate = float(m_tail.group(4).replace(",", ""))
                    full_desc = (pending_desc + " " + extra_desc).strip()
                    items.append((full_desc, int(qty) if qty.is_integer() else qty, unit, rate))
                    last_idx = len(items) - 1
                    pending_desc = None
                    continue

                pending_desc = (pending_desc + " " + line.strip()).strip()
                continue

            match = parse_item_line(line)
            if not match and i < len(lines) - 1:
                combined = f"{line} {lines[i + 1]}"
                match = parse_item_line(combined)
                if match:
                    skip_next = True

            if match:
                desc, qty, unit, rate = match
                items.append((desc, int(qty) if qty.is_integer() else qty, unit, rate))
                last_idx = len(items) - 1
                pending_desc = None
            elif last_idx is not None:
                if not re.search(r"^\d+(?:\.\d+)?\.?\s+", line):
                    if re.search(r"\b(TOTAL|SST|TERM|PAYMENT|DELIVERY|VALIDITY|WARRANTY|BANK|ACCOUNT|REFERENCE|QUOTE|QUOTATION)\b", upper):
                        in_items = False
                        last_idx = None
                        continue
                    prev = items[last_idx]
                    items[last_idx] = (prev[0] + " " + line.strip(), prev[1], prev[2], prev[3])
            else:
                m_desc_only = re.search(r"^\s*\d+(?:\.\d+)?\.?\s+(.+)$", line)
                if m_desc_only:
                    pending_desc = m_desc_only.group(1).strip().rstrip(".")

        return items

    def extract_items_from_tables(self, pdf):
        items = []
        table_settings = {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "intersection_tolerance": 5,
            "snap_tolerance": 3,
            "join_tolerance": 3,
            "edge_min_length": 3,
        }

        def normalize(s):
            return re.sub(r"\s+", " ", (s or "")).strip().lower()

        def parse_number(val):
            if val is None:
                return None
            txt = normalize(val).replace("rm", "").replace(",", "").strip()
            m = re.search(r"(\d+(?:\.\d{2})?)", txt)
            return float(m.group(1)) if m else None

        last_idx = None
        for page in pdf.pages:
            try:
                tables = page.extract_tables(table_settings) or []
            except Exception:
                tables = []
            for table in tables:
                if not table or len(table) < 2:
                    continue
                header = [normalize(c) for c in table[0]]
                desc_idx = None
                qty_idx = None
                unit_price_idx = None

                for i, h in enumerate(header):
                    if "description" in h or "product details" in h:
                        desc_idx = i
                    if h in ("qty", "quantity"):
                        qty_idx = i
                    if "unit price" in h or "price ex.sst" in h or "price ex" in h:
                        unit_price_idx = i
                    if "unit price (rm)" in h or h == "unit price":
                        unit_price_idx = i

                if desc_idx is None:
                    continue

                for row in table[1:]:
                    if not row or all((c or "").strip() == "" for c in row):
                        continue
                    row_norm = [normalize(c) for c in row]
                    if any("description" in c for c in row_norm):
                        continue

                    desc = row[desc_idx] if desc_idx < len(row) else ""
                    qty_val = row[qty_idx] if (qty_idx is not None and qty_idx < len(row)) else None
                    price_val = row[unit_price_idx] if (unit_price_idx is not None and unit_price_idx < len(row)) else None

                    qty = parse_number(qty_val) or None
                    rate = parse_number(price_val) or None

                    if desc and (qty is not None or rate is not None):
                        items.append((desc.strip().rstrip("."), int(qty or 1), "Unit", float(rate or 0)))
                        last_idx = len(items) - 1
                    elif desc and last_idx is not None:
                        prev = items[last_idx]
                        items[last_idx] = (prev[0] + " " + desc.strip(), prev[1], prev[2], prev[3])

        return items

    def fallback_total_item(self, text):
        price_matches = re.findall(r"(?:Grand\s*)?Total.*?([\d,]+\.\d{2})", text, re.IGNORECASE)
        total_price = float(price_matches[-1].replace(",", "")) if price_matches else 0.00
        return ("Extracted Total (Please verify PDF for line items)", 1, "Lot", total_price)

    # ---------------- Excel helpers ----------------
    def get_vendor_columns(self, ws, extracted_vendor_name):
        vendor_slots = [8, 10, 12]
        for col in vendor_slots:
            cell_val = str(ws.cell(row=6, column=col).value or "").strip()
            if cell_val:
                if extracted_vendor_name.lower() in cell_val.lower() or cell_val.lower() in extracted_vendor_name.lower():
                    return col, col + 1

        for col in vendor_slots:
            if not ws.cell(row=6, column=col).value:
                print(f"Assigning new vendor '{extracted_vendor_name}' to Column {col}")
                ws.cell(row=6, column=col).value = extracted_vendor_name
                return col, col + 1

        return 8, 9

    def find_footer_start(self, ws):
        for r in range(8, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                if val in self.FOOTER_LABELS:
                    return r
        return 16

    def find_label_row(self, ws, label_text):
        target = label_text.strip().lower()
        for r in range(8, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                if val == target:
                    return r
        return None

    def copy_item_row_styles(self, ws, insert_start_row, rows_to_insert):
        template_rows = [8, 9, 10]
        max_col = ws.max_column
        for offset in range(rows_to_insert):
            target_row = insert_start_row + offset
            source_row = template_rows[offset % len(template_rows)]
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
            for c in range(1, max_col + 1):
                src = ws.cell(row=source_row, column=c)
                tgt = ws.cell(row=target_row, column=c)
                if src.has_style:
                    tgt._style = copy(src._style)
            self.apply_item_row_merges(ws, target_row)

    def apply_item_row_merges(self, ws, target_row):
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row == 8 and rng.max_row == 8:
                try:
                    ws.merge_cells(
                        start_row=target_row,
                        end_row=target_row,
                        start_column=rng.min_col,
                        end_column=rng.max_col,
                    )
                except ValueError:
                    pass


if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = QCFApp(root)
    root.mainloop()
